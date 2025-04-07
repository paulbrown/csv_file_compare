import csv
import chardet
from dictdiffer import diff
import typer
from pathlib import Path
from typing_extensions import Annotated
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Function to import CSV and zip heading and lines into a data structure
def load_csv(filePath: str, keys: list[str]):
    keySeperator: str = "_"

    with open(filePath, mode = 'rb') as f:
        data = f.read()
        encoding = chardet.detect(data).get("encoding")
        print(f"Encoding for {filePath} has been determined: {encoding}")
    
        with open(filePath, mode = 'r', encoding = encoding) as file:          
            csvFile = csv.reader(file)
            line1 = next(csvFile)

            # HACKATY-HACKATY YAK YAK: Matt made me do it!!!
            # Make the bad row1 in exported SF report row go-away
            # if present in CSV, use the correct line for header zipping
            headerCrap = ["Exported", "Exportado", "diekspor", "Exporté", "Eksportert", "Wyeksportowano", "Exporterat", "Экспортировано", "导出到"]
            if any(item in headerCrap for item in line1[0].split()):
                next(csvFile)
                heading = next(csvFile)
            else:
                heading = line1

            rows = [dict(zip(heading, line)) for line in csvFile]
            keyFn = lambda r: keySeperator.join([r[x] for x in keys])
            # Bit of dictionary comprehension to make our key using lambda function
            result = {keyFn(r): r for r in rows}
            return result
    

# Function to compare 2 CSV file which have the same nature or structure
def compare(previous: dict[str, dict[str, str]], current: dict[str, dict[str, str]]):
    result = {
        "deleted": [],
        "inserted": [],
        "updated": []
    }

    # Have any key rows been removed
    deleted = [id for id in previous if id not in current]
    # Have any key rows been added
    inserted = [id for id in current if id not in previous]
    # Unique list of key identifiers 
    deleted_inserted = set(deleted) | set(inserted)
    # Remaining key rows not part of added or removed but could have changed
    potential_update = [id for id in current if id not in deleted_inserted]
    updated = [id for id in potential_update if current[id] != previous[id]]

    # Build dictionary for removed records
    if deleted:
        for id in deleted:
            # Merge our deleted chucks into 1 dictionary
            delete = {"key": id} | previous[id]
            result["deleted"].append(delete)

    # Build dictionary for added records 
    if inserted:
        for id in inserted:
            # Merge our inserted chucks into 1 dictionary
            insert = {"key": id} | current[id]
            result["inserted"].append(insert)

    # Build dictionary for updated records 
    if updated:
        for id in updated:
            # Used established dictdiffer library to find differences
            diffs = list(diff(previous[id], current[id]))
            if diffs:
                # Some comprehension to produce as single list of key, value 
                # pairs for each record (row by key) that has had updates 
                updates = {
                    "updates": {
                        field[0] if isinstance(field, list) else field: [previous_value, current_value] 
                            for _, field, (previous_value, current_value) in diffs
                    }
                }
                # Merge our updated chucks into 1 dictionary
                update = {"key": id} | current[id] | updates
                result["updated"].append(update)
              
    return result


# Function to ouput result into seperate DELETED, INSERTED, UPDATED XLSX sheets
def build_xlsx_output(output_xlsx: str, result: dict[str, list]):
    # Do some list comprehension because I wanted to see if I still could!!!
    # Converts nested dictionary to a list of changes (tuples) for each key
    # (row, column, old_value, new_value)
    updates = [
        (value, field_chg, value_chg[0], value_chg[1]) 
            for updated in result["updated"]
            for (key, value) in updated.items() if key == "key"
            for (field_chg, value_chg) in updated["updates"].items()
    ]
    
    # Write output to Excel
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:

        if result["deleted"]:
            # Output to Excel rows inserted into current file
            dfDeleted = pd.DataFrame(result["deleted"]).set_index("key")
            dfDeleted.to_excel(writer, index=True, sheet_name="DELETED")
        else: 
            dfDeleted = pd.DataFrame({"key": ["Nothing to be deleted :)"]})
            dfDeleted.to_excel(writer, index=False, sheet_name="DELETED")

        if result["inserted"]:
            # Output to Excel rows inserted into current file
            dfInserted = pd.DataFrame(result["inserted"]).set_index("key")
            dfInserted.to_excel(writer, index=True, sheet_name="INSERTED")
        else:
            dfInserted = pd.DataFrame({"key": ["Nothing to be inserted :)"]})
            dfInserted.to_excel(writer, index=False, sheet_name="INSERTED") 

        if result["updated"]:
            # Outpus to Excel rows updated between files
            dfUpdated = pd.DataFrame(result["updated"]).set_index("key")
            dfUpdated.to_excel(writer, index=True, sheet_name="UPDATED")
            
            # Open workbook for highlighting
            wb = writer.book
            ws = wb['UPDATED']

            # Highlight cell where change occured
            highlight = PatternFill(patternType="solid", fgColor="FFFF00")
            for update in updates:
                row_index = dfUpdated.index.get_loc(update[0])
                column_index = dfUpdated.columns.get_loc(update[1])
                new_value = dfUpdated.iat[row_index, column_index]
                # Check this is in fact the data/cell that has updated before highlighting
                # update[3] = current value 
                # update[2] previous value
                if new_value == update[3] and new_value != update[2]:
                    # Add offset of 2 in each direction just because 
                    # it's Excel and we need the yellow in the right place 
                    ws.cell(row_index + 2, column_index + 2).fill = highlight  
        else:
            dfUpdated = pd.DataFrame({"key": ["Nothing to be updated :)"]})
            dfUpdated.to_excel(writer, index=False, sheet_name="UPDATED")

# Main entry-point for csv-file-compare utility
def main(previous_csv_filepath: Annotated[Path, typer.Option(exists=True, file_okay=True, dir_okay=False, writable=False, readable=True, resolve_path=True)],
         current_csv_filepath: Annotated[Path, typer.Option(exists=True, file_okay=True, dir_okay=False, writable=False, readable=True, resolve_path=True)],
         output_xlsx_filepath: Annotated[Path, typer.Option(exists=False, file_okay=True, dir_okay=False, writable=True, readable=False, resolve_path=True)] = "Differences.xlsx",
         key: Annotated[Optional[List[str]], typer.Option()] = None):

    # Check key option has been provided
    if not key:
        print(f"No provided --key 'field name' option supplied.")
        raise typer.Abort()
    else: 
        print(f'Processing file differences based on input key/s: {", ".join(str(i) for i in key)}')

    # Check previous .csv file has been supplied to compare
    if previous_csv_filepath is None:
        print("No previous CSV file")
        raise typer.Abort()
    if previous_csv_filepath.is_file():
        previous_csv = load_csv(previous_csv_filepath, key)
        print(f"Your initial baseline CSV file has been read from: {previous_csv_filepath}")
    elif previous_csv_filepath.is_dir():
        print("File path is a directory, please supply full path to previous .CSV file")
    elif not previous_csv_filepath.exists():
        print("The previous CSV file doesn't exist")

    # Check current .csv file has been supplied to compare
    if current_csv_filepath is None:
        print("No current CSV file")
        raise typer.Abort()
    if current_csv_filepath.is_file():
        current_csv = load_csv(current_csv_filepath, key)
        print(f"Your comparison CSV file has been read from: {current_csv_filepath}")
    elif current_csv_filepath.is_dir():
        print("File path is a directory, please supply full path to current .CSV file")
    elif not current_csv_filepath.exists():
        print("The current CSV file doesn't exist")

    # Check if output .xlsx file has been supplied
    if output_xlsx_filepath.is_file():
        print(f"Your existing output XLSX file is: {output_xlsx_filepath}")
    elif output_xlsx_filepath.is_dir():
        print("File path is a directory, please supply full path to output .XLSX file")
    else: print(f"Your output XLSX file will be written to: {output_xlsx_filepath}")

    # Return results following compare between pervious and current files
    diffs = compare(previous_csv, current_csv)
 
    # Build .xlsx output file
    build_xlsx_output(output_xlsx_filepath, diffs)

    # add better help
    # package up into single executable

if __name__ == "__main__":
    typer.run(main)